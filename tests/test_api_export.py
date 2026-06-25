"""XLSX + CSV export tests.

We exercise the produced file end-to-end with openpyxl rather than
mocking — the test is about whether the produced artefact is correct,
not about openpyxl call shape.
"""

from __future__ import annotations

import csv as csv_mod
from datetime import UTC, date, datetime
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.api.days_service import AdjustParams, adjust_day, lock_day
from app.exporters.csv import write_csv
from app.exporters.xlsx import SUMMARY_HEADERS, write_xlsx
from app.models import Observation
from app.sessions.persistence import sessionise_date
from app.sessions.rules import RuleSet


def utc(y: int, m: int, d: int, hh: int = 0, mm: int = 0, ss: int = 0) -> datetime:
    return datetime(y, m, d, hh, mm, ss, tzinfo=UTC)


@pytest.fixture
def seeded_year(db_session: Session) -> Session:
    """Two seeded days in FY 2025-26, one locked one with an adjustment."""
    from app.config import get_settings
    from app.main import seed_config_if_missing

    seed_config_if_missing(db_session, get_settings())
    db_session.commit()
    rules = RuleSet.from_db(db_session)

    # Day 1: 2025-08-15 (FY 2025-26), 3 hours, locked.
    # Sydney in August is AEST (UTC+10): 09:00 AEST == 23:00 UTC previous day.
    for ts, conn in [
        (utc(2025, 8, 14, 23, 0), True),  # 09:00 Sydney AEST on the 15th
        (utc(2025, 8, 15, 2, 0), False),  # 12:00 Sydney AEST on the 15th
    ]:
        db_session.add(
            Observation(
                observed_at=ts,
                controller_seen_at=ts,
                mac="a",
                device_label="iPhone",
                ssid="WFH-TEST",
                is_connected=conn,
                signal_dbm=None,
                raw_json="{}",
            )
        )
    db_session.commit()
    sessionise_date(db_session, date(2025, 8, 15), rules)
    db_session.commit()
    adjust_day(
        db_session,
        date(2025, 8, 15),
        AdjustParams(adjustment_seconds=-45 * 60, reason="lunch"),
    )
    db_session.commit()
    lock_day(db_session, date(2025, 8, 15))
    db_session.commit()

    # Day 2: 2026-02-03 (FY 2025-26), 4 hours, unlocked.
    # Sydney in February is AEDT (UTC+11): 09:00 AEDT == 22:00 UTC previous day.
    for ts, conn in [
        (utc(2026, 2, 2, 22, 0), True),  # 09:00 Sydney AEDT on the 3rd
        (utc(2026, 2, 3, 2, 0), False),  # 13:00 Sydney AEDT on the 3rd
    ]:
        db_session.add(
            Observation(
                observed_at=ts,
                controller_seen_at=ts,
                mac="a",
                device_label="iPhone",
                ssid="WFH-TEST",
                is_connected=conn,
                signal_dbm=None,
                raw_json="{}",
            )
        )
    db_session.commit()
    sessionise_date(db_session, date(2026, 2, 3), rules)
    db_session.commit()

    return db_session


class TestXlsxStructure:
    def test_workbook_has_three_named_sheets(self, seeded_year: Session) -> None:
        buf = BytesIO()
        write_xlsx(seeded_year, date(2025, 7, 1), date(2026, 6, 30), "2025-26", buf)
        buf.seek(0)
        wb = load_workbook(buf)
        assert wb.sheetnames == ["Summary", "Year total", "Methodology"]

    def test_summary_headers_are_complete(self, seeded_year: Session) -> None:
        buf = BytesIO()
        write_xlsx(seeded_year, date(2025, 7, 1), date(2026, 6, 30), "2025-26", buf)
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb["Summary"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(SUMMARY_HEADERS))]
        assert headers == SUMMARY_HEADERS

    def test_summary_has_one_row_per_seeded_date(self, seeded_year: Session) -> None:
        buf = BytesIO()
        write_xlsx(seeded_year, date(2025, 7, 1), date(2026, 6, 30), "2025-26", buf)
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb["Summary"]
        # Two seeded days, header row at row 1.
        data_dates = []
        r = 2
        while ws.cell(row=r, column=1).value is not None:
            data_dates.append(ws.cell(row=r, column=1).value)
            r += 1
        assert len(data_dates) == 2

    def test_year_total_has_summary_cells(self, seeded_year: Session) -> None:
        buf = BytesIO()
        write_xlsx(seeded_year, date(2025, 7, 1), date(2026, 6, 30), "2025-26", buf)
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb["Year total"]
        # Day 1: 3h - 0.75 = 2.25h. Day 2: 4h. Total = 6.25h.
        assert ws["B3"].value is not None
        assert abs(float(ws["B3"].value) - 6.25) < 0.001
        # Locked / unlocked counts.
        assert ws["B4"].value == 1  # locked
        assert ws["B5"].value == 1  # unlocked
        # Fixed-rate cell is blank with a comment.
        assert ws["B8"].value is None
        assert ws["B8"].comment is not None
        # Formula computes A*B.
        assert ws["B9"].value == "=B3*B8"

    def test_methodology_sheet_has_placeholders_replaced(self, seeded_year: Session) -> None:
        buf = BytesIO()
        write_xlsx(seeded_year, date(2025, 7, 1), date(2026, 6, 30), "2025-26", buf)
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb["Methodology"]
        # Collect all non-empty cell text.
        all_text = "\n".join(ws.cell(row=r, column=1).value or "" for r in range(1, ws.max_row + 1))
        # WORK_SSID should be substituted from the seeded config (WFH-TEST).
        assert "WFH-TEST" in all_text
        # Default-annotated GAP_BRIDGE_MINUTES → 10.
        assert "10" in all_text
        # FY label substituted.
        assert "2025-26" in all_text
        # Original placeholders should be gone.
        assert "[WORK_SSID]" not in all_text
        assert "[GAP_BRIDGE_MINUTES, default 10]" not in all_text

    def test_locked_at_renders_as_datetime(self, seeded_year: Session) -> None:
        buf = BytesIO()
        write_xlsx(seeded_year, date(2025, 7, 1), date(2026, 6, 30), "2025-26", buf)
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb["Summary"]
        # The first data row is the locked day (2025-08-15 was locked).
        locked_at = ws.cell(row=2, column=9).value
        assert locked_at is not None
        assert isinstance(locked_at, datetime)


class TestCsvOutput:
    def test_csv_headers_and_rows(self, seeded_year: Session, tmp_path) -> None:  # type: ignore[no-untyped-def]
        out = tmp_path / "out.csv"
        n = write_csv(seeded_year, date(2025, 7, 1), date(2026, 6, 30), out, "Australia/Sydney")
        assert n == 2
        with out.open(encoding="utf-8") as f:
            reader = csv_mod.reader(f)
            rows = list(reader)
        assert rows[0][0] == "date"
        assert len(rows) == 3  # header + 2 data rows


class TestExportEndpoints:
    def test_xlsx_endpoint(self, seeded_year: Session, client: TestClient) -> None:
        # `client` is from conftest; uses a fresh DB. seeded_year used a
        # *different* session over the same DB, so the data IS persisted.
        # seeded_year has one unlocked day (2026-02-03), so the Phase 10.E
        # export guard requires allow_unlocked=true to produce the file.
        resp = client.get(
            "/api/export.xlsx", params={"fy": "2025-26", "allow_unlocked": "true"}
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument"
        )
        assert b"wfh-logbook-2025-26.xlsx" in resp.headers["content-disposition"].encode()
        # Should be a real xlsx (zip magic PK).
        assert resp.content[:2] == b"PK"

    def test_csv_endpoint(self, seeded_year: Session, client: TestClient) -> None:
        resp = client.get(
            "/api/export.csv",
            params={"from": "2025-07-01", "to": "2026-06-30"},
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("text/csv")
        first_line = resp.text.splitlines()[0]
        assert first_line.startswith("date,")

    def test_bad_fy_400(self, client: TestClient) -> None:
        resp = client.get("/api/export.xlsx", params={"fy": "not-an-fy"})
        assert resp.status_code == 400
