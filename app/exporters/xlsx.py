"""XLSX exporter (the tax-filing deliverable).

Three sheets per HANDOFF §6 Phase 5:

- **Summary**: one row per day in the FY, with computed/adjustment/claimed
  hours, reason, version, locked status, rule_version.
- **Year total**: aggregates + a fixed-rate cell (left blank with a comment
  per HANDOFF §6 Phase 5 and METHODOLOGY §7) and a formula computing the
  dollar figure.
- **Methodology**: a copy of docs/METHODOLOGY.md with bracketed placeholders
  populated from the live Config + Devices rows.
"""

from __future__ import annotations

import logging
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.exporters.common import (
    build_summary_rows,
    get_active_devices,
    get_config,
    render_methodology_with_config,
)

if TYPE_CHECKING:
    from sqlalchemy.orm import Session


logger = logging.getLogger(__name__)


# Resolved at module-import; tests don't override.
METHODOLOGY_MD_PATH = Path(__file__).resolve().parents[2] / "docs" / "METHODOLOGY.md"


SUMMARY_HEADERS = [
    "Date",
    "Day of week",
    "Computed hours",
    "Adjustment (hours)",
    "Adjustment reason",
    "Claimed hours",
    "Version",
    "Locked",
    "Locked at",
    "Rule version",
    "Recorded via",
]


def _write_summary_sheet(
    wb: Workbook,
    db: Session,
    from_date: date,
    to_date: date,
    tz_name: str,
) -> int:
    """Returns the count of data rows written."""
    ws = wb.active
    assert ws is not None
    ws.title = "Summary"
    header_font = Font(bold=True)
    for col, name in enumerate(SUMMARY_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = header_font
        c.alignment = Alignment(horizontal="left")

    rows = build_summary_rows(db, from_date, to_date, tz_name)
    for r, row in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=row.local_date)
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=row.day_of_week)
        ws.cell(row=r, column=3, value=round(row.computed_hours, 4))
        ws.cell(row=r, column=4, value=round(row.adjustment_hours, 4))
        ws.cell(row=r, column=5, value=row.adjustment_reason)
        ws.cell(row=r, column=6, value=round(row.claimed_hours, 4))
        ws.cell(row=r, column=7, value=row.version)
        ws.cell(row=r, column=8, value="Yes" if row.locked else "No")
        if row.locked_at is not None:
            ws.cell(row=r, column=9, value=row.locked_at.replace(tzinfo=None))
            ws.cell(row=r, column=9).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=r, column=10, value=row.rule_version)
        ws.cell(row=r, column=11, value=row.created_by)

    # Column widths.
    widths = [12, 12, 14, 16, 50, 14, 8, 8, 18, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze the header row.
    ws.freeze_panes = "A2"
    return len(rows)


def _write_year_total_sheet(
    wb: Workbook,
    db: Session,
    from_date: date,
    to_date: date,
    tz_name: str,
    fy_label: str,
) -> None:
    ws = wb.create_sheet("Year total")
    rows = build_summary_rows(db, from_date, to_date, tz_name)
    cfg = get_config(db)
    daily_cap_seconds = cfg.daily_cap_hours * 3600

    total_claimed = sum(r.claimed_hours for r in rows)
    locked_days = sum(1 for r in rows if r.locked)
    unlocked_days = sum(1 for r in rows if not r.locked)
    anomalous_days = sum(1 for r in rows if r.claimed_hours * 3600 > daily_cap_seconds)

    header_font = Font(bold=True, size=12)
    label_font = Font(bold=True)
    accent_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")

    ws.cell(row=1, column=1, value=f"WFH Logbook — Year total — FY {fy_label}").font = header_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    ws.cell(row=3, column=1, value="Total claimed hours").font = label_font
    ws.cell(row=3, column=2, value=round(total_claimed, 2))

    ws.cell(row=4, column=1, value="Locked days").font = label_font
    ws.cell(row=4, column=2, value=locked_days)

    ws.cell(row=5, column=1, value="Unlocked days").font = label_font
    ws.cell(row=5, column=2, value=unlocked_days)

    ws.cell(row=6, column=1, value="Anomalous days (>daily cap)").font = label_font
    ws.cell(row=6, column=2, value=anomalous_days)

    ws.cell(row=8, column=1, value="ATO fixed-rate (per hour)").font = label_font
    rate_cell = ws.cell(row=8, column=2, value=None)
    rate_cell.fill = accent_fill
    rate_cell.comment = Comment(
        "Set this to the ATO published rate for the relevant year. "
        "This export does not assert a dollar value — see METHODOLOGY §7.",
        "WFH Logbook",
    )
    rate_cell.number_format = "$#,##0.00"

    ws.cell(row=9, column=1, value="Fixed-rate $ (computed)").font = label_font
    formula_cell = ws.cell(row=9, column=2, value="=B3*B8")
    formula_cell.number_format = "$#,##0.00"

    ws.cell(row=11, column=1, value="Configuration snapshot").font = label_font
    snapshot_rows = [
        ("Work SSID", cfg.work_ssid),
        ("Gap-bridge minutes", cfg.gap_bridge_minutes),
        ("Minimum session minutes", cfg.min_session_minutes),
        ("Daily cap hours (review flag)", cfg.daily_cap_hours),
        ("Local timezone", cfg.local_timezone),
        ("Rule version", cfg.rule_version),
    ]
    for i, (label, val) in enumerate(snapshot_rows, start=12):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    devices = get_active_devices(db)
    if devices:
        dev_start = 12 + len(snapshot_rows)
        ws.cell(row=dev_start, column=1, value="Devices tracked").font = label_font
        for i, d in enumerate(devices, start=dev_start + 1):
            ws.cell(row=i, column=1, value=d.label)
            ws.cell(row=i, column=2, value=d.mac)

    # Disclaimer.
    disc_row = ws.max_row + 2
    disc = ws.cell(
        row=disc_row,
        column=1,
        value=(
            "Hours only. This export does not assert eligibility, the appropriate "
            "method to use, or a dollar value. See METHODOLOGY §7. "
            "Records must be kept for 5 years from lodgement (PCG 2023/1)."
        ),
    )
    disc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=disc_row, start_column=1, end_row=disc_row, end_column=2)
    ws.row_dimensions[disc_row].height = 60

    # Column widths.
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22


def _write_methodology_sheet(
    wb: Workbook,
    db: Session,
    fy_label: str,
    methodology_text: str | None = None,
) -> None:
    """Embed METHODOLOGY.md with placeholders replaced.

    If ``methodology_text`` is None, read from the on-disk template.
    """
    ws = wb.create_sheet("Methodology")
    cfg = get_config(db)
    if methodology_text is None:
        try:
            methodology_text = METHODOLOGY_MD_PATH.read_text(encoding="utf-8")
        except OSError:
            logger.warning("methodology template not found at %s", METHODOLOGY_MD_PATH)
            methodology_text = "(methodology template not available at export time)"

    text = render_methodology_with_config(methodology_text, cfg, fy_label)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)

    for i, line in enumerate(text.splitlines(), start=1):
        cell = ws.cell(row=i, column=1, value=line)
        # Light formatting for headings.
        if line.startswith("# "):
            cell.font = title_font
        elif line.startswith(("## ", "### ")):
            cell.font = section_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110


def write_xlsx(
    db: Session,
    from_date: date,
    to_date: date,
    fy_label: str,
    out: Path | BytesIO,
    tz_name: str | None = None,
) -> int:
    """Write the XLSX file/buffer. Returns the number of data rows in Summary."""
    cfg = get_config(db)
    tz = tz_name or cfg.local_timezone

    wb = Workbook()
    rows_written = _write_summary_sheet(wb, db, from_date, to_date, tz)
    _write_year_total_sheet(wb, db, from_date, to_date, tz, fy_label)
    _write_methodology_sheet(wb, db, fy_label)

    if isinstance(out, BytesIO):
        wb.save(out)
    else:
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
    logger.info(
        "xlsx export: fy=%s from=%s to=%s rows=%d out=%s",
        fy_label,
        from_date,
        to_date,
        rows_written,
        out,
    )
    return rows_written
